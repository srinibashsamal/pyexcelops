#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Program Name    : excel_ops.py
Description     : A professional styling tool for Excel. It takes one or more
                  DataFrames and writes them into a single, consistently
                  formatted .xlsx workbook: bold white headers on a navy fill,
                  thin borders, best-fit column widths, wrapped text, and
                  optional native Excel tables and filters, so the file is
                  ready to read for management without any manual clean-up.

Author          : Srinibash Samal
"""

import math
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Union

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =====================================================================
# SECTION 1: CONSTANTS
# =====================================================================
_LIGHT_BLUE: str = "ADCFEB"
_NAVY_BLUE: str = "002266"
_WHITE: str = "FFFFFF"
_FONT_NAME: str = "Calibri"
_DEFAULT_TABLE_STYLE: str = "TableStyleMedium1"
_NO_DATA_MESSAGE: str = "This sheet contains no data based on the current inputs!"

# Extra width added so the filter dropdown icon does not cover the header text
_FILTER_ICON_WIDTH: float = 3.0

# Excel's own default column width, used when a width was never set explicitly
_DEFAULT_COL_WIDTH: float = 8.43


# =====================================================================
# SECTION 2: STYLE AND SIZING HELPERS
# =====================================================================
def _make_styles() -> Dict[str, Any]:
    """Build the reusable openpyxl style objects once, to share across all sheets."""
    thin = Side(style="thin")

    return {
        "header_font": Font(name=_FONT_NAME, bold=True, color=_WHITE),
        "header_fill": PatternFill(
            start_color=_NAVY_BLUE, end_color=_NAVY_BLUE, fill_type="solid"
        ),
        "highlight_fill": PatternFill(
            start_color=_LIGHT_BLUE, end_color=_LIGHT_BLUE, fill_type="solid"
        ),
        "header_align": Alignment(
            horizontal="center", vertical="center", wrap_text=True
        ),
        "body_wrap_align": Alignment(vertical="top", wrap_text=True),
        "thin_border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _calc_column_width(
    series: pd.Series,
    header: str,
    padding: int = 2,
    max_width: int = 50,
) -> int:
    """Return a best-fit column width in Excel character units, capped at `max_width`.

    The width is driven by the longest value in the column or the header,
    whichever is wider.
    """
    if series.empty:
        return min(len(header) + padding, max_width)

    max_len = max(series.astype(str).str.len().max(), len(header))
    return min(max_len + padding, max_width)


def _apply_wrap_row_heights(
    ws,
    min_row: int,
    max_row: int,
    wrap_cols: List[int],
    base_line_height: float = 15.0,
    min_height: float = 15.0,
    max_height: float = 240.0,
) -> None:
    """
    Heuristically set row heights for rows that contain wrapped text.

    Estimates the number of visible lines per cell from the column width and
    the length of the cell value, then sets the row height accordingly.
    """
    if not wrap_cols:
        return

    # Cache the width of each wrapping column up front
    col_widths: Dict[int, float] = {}
    for col_idx in wrap_cols:
        raw_width = ws.column_dimensions[get_column_letter(col_idx)].width
        col_widths[col_idx] = float(raw_width) if raw_width else _DEFAULT_COL_WIDTH

    fudge = 1.05  # small guard against under-estimating the line count

    for row_num in range(min_row, max_row + 1):
        needed_lines = 1

        for col_idx in wrap_cols:
            cell = ws.cell(row=row_num, column=col_idx)
            text = str(cell.value) if cell.value is not None else ""
            if not text:
                continue

            chars_per_line = max(int(col_widths[col_idx] / fudge), 5)

            # Count each manual line break separately, then wrap within each one
            lines_for_cell = sum(
                max(1, math.ceil(len(para) / chars_per_line)) if para else 1
                for para in text.split("\n")
            )
            needed_lines = max(needed_lines, lines_for_cell)

        # The tallest cell in the row decides the row height, within the limits
        ws.row_dimensions[row_num].height = max(
            min_height, min(max_height, needed_lines * base_line_height)
        )


# =====================================================================
# SECTION 3: TABLE NAMING HELPERS
# =====================================================================
def _sanitize_table_name(name: str) -> str:
    """Turn any string into a valid Excel table `displayName`.

    Excel requires no spaces or symbols, a leading letter or underscore,
    and a maximum of 255 characters.
    """
    base = re.sub(r"[^A-Za-z0-9_]", "_", name.strip()) or "Table"

    if not re.match(r"^[A-Za-z_]", base):
        base = "_" + base

    return base[:255]


def _unique_table_name(workbook, desired: str) -> str:
    """Return `desired`, or `desired_2`, `desired_3`... if the name is already taken.

    Table names must be unique across the entire workbook, not just per sheet.
    """
    existing: Set[str] = set()
    for worksheet in workbook.worksheets:
        existing.update(worksheet.tables.keys())

    if desired not in existing:
        return desired

    suffix = 2
    while f"{desired}_{suffix}" in existing:
        suffix += 1

    return f"{desired}_{suffix}"


# =====================================================================
# SECTION 4: VALIDATORS
# =====================================================================
def _ensure_sheet_names(
    dfs: List[pd.DataFrame], sheet_names: Optional[List[str]]
) -> List[str]:
    """
    Ensure sheet_names list matches the number of DataFrames.
    Fills missing names with 'Sheet1', 'Sheet2', etc.
    """
    final_names = list(sheet_names or [])

    for i in range(len(final_names), len(dfs)):
        final_names.append(f"Sheet{i + 1}")

    return final_names[: len(dfs)]


def _validate_highlight_cols(
    dfs: List[pd.DataFrame],
    highlight_cols_list: Optional[List[Optional[List[str]]]],
) -> List[Optional[List[str]]]:
    """Check the highlight list lines up with the DataFrames, one entry per sheet."""
    if highlight_cols_list is None:
        return [None] * len(dfs)

    if len(highlight_cols_list) != len(dfs):
        raise ValueError(
            f"'highlight_cols_list' length ({len(highlight_cols_list)}) "
            f"must match 'dfs' length ({len(dfs)})."
        )

    return highlight_cols_list


# =====================================================================
# SECTION 5: SHEET FORMATTING STEPS
# =====================================================================
def _write_empty_sheet(
    ws, ncols: int, styles: Dict[str, Any], wrap_threshold: int
) -> None:
    """Write a friendly "no data" notice on a sheet that has headers but no rows."""
    # No columns at all: just drop the message in the first cell
    if ncols == 0:
        ws["A1"] = _NO_DATA_MESSAGE
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(name=_FONT_NAME, italic=True, color=_NAVY_BLUE)
        return

    # Otherwise centre the message across the header row's full width
    msg_cell = ws.cell(row=2, column=1)
    msg_cell.value = _NO_DATA_MESSAGE
    msg_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    msg_cell.font = Font(name=_FONT_NAME, italic=True, color=_NAVY_BLUE)

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ws.row_dimensions[2].height = 20

    for row in ws.iter_rows(min_row=1, max_row=2, max_col=ncols):
        for cell in row:
            cell.border = styles["thin_border"]

    for col_idx, header_cell in enumerate(ws[1], start=1):
        width = min(len(str(header_cell.value)) + 2, wrap_threshold * 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _format_headers_and_widths(
    ws,
    df: pd.DataFrame,
    styles: Dict[str, Any],
    wrap_threshold: int,
    filter_padding: float,
) -> List[int]:
    """Style the header row, size every column, and wrap the wide ones.

    Returns:
        Indices of the columns that were set to wrap, so row heights can be
        estimated for them later.
    """
    wrapped_col_indices: List[int] = []
    nrows = len(df)

    for col_idx, header_cell in enumerate(ws[1], start=1):
        header_cell.font = styles["header_font"]
        header_cell.fill = styles["header_fill"]
        header_cell.alignment = styles["header_align"]
        header_cell.border = styles["thin_border"]

        width = _calc_column_width(
            df.iloc[:, col_idx - 1],
            str(header_cell.value),
            max_width=wrap_threshold * 2,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width + filter_padding

        # Anything wider than the threshold reads better wrapped than stretched
        if width >= wrap_threshold:
            wrapped_col_indices.append(col_idx)
            for row in ws.iter_rows(
                min_row=2, max_row=nrows + 1, min_col=col_idx, max_col=col_idx
            ):
                row[0].alignment = styles["body_wrap_align"]

    return wrapped_col_indices


def _apply_borders(ws, nrows: int, ncols: int, styles: Dict[str, Any]) -> None:
    """Draw thin borders across the header row and all data rows."""
    for row in ws.iter_rows(min_row=1, max_row=nrows + 1, max_col=ncols):
        for cell in row:
            cell.border = styles["thin_border"]


def _highlight_columns(
    ws, nrows: int, highlight_cols: Optional[List[str]], styles: Dict[str, Any]
) -> None:
    """Give the data cells of the named columns the light-blue highlight fill."""
    highlight_set = set(highlight_cols or [])
    if not highlight_set:
        return

    for col_idx, header in enumerate([cell.value for cell in ws[1]], start=1):
        if header in highlight_set:
            for row_num in range(2, nrows + 2):
                ws.cell(row=row_num, column=col_idx).fill = styles["highlight_fill"]


def _add_table_or_filter(
    ws,
    workbook,
    sheet_name: str,
    nrows: int,
    ncols: int,
    as_table: bool,
    table_style: str,
    add_filters: bool,
) -> None:
    """Convert the range to a native Excel table, or just add filter dropdowns."""
    data_range = f"A1:{get_column_letter(ncols)}{nrows + 1}"

    if as_table:
        # Table names must be unique workbook-wide, so resolve any clash first
        desired_name = _sanitize_table_name(f"{sheet_name}_Table")
        table_name = _unique_table_name(workbook, desired_name)

        table = Table(displayName=table_name, ref=data_range)
        table.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    elif add_filters:
        # Tables bring their own filters, so this only applies without one
        ws.auto_filter.ref = data_range


# =====================================================================
# SECTION 6: PUBLIC API
# =====================================================================
def save_formatted_excel(
    dfs: List[pd.DataFrame],
    output_path: Union[str, Path],
    sheet_names: Optional[List[str]] = None,
    highlight_cols_list: Optional[List[Optional[List[str]]]] = None,
    freeze_header: bool = False,
    wrap_threshold: int = 40,
    as_table: bool = False,
    table_style: str = _DEFAULT_TABLE_STYLE,
    auto_fit_row_heights: bool = False,
    show_grids: bool = True,
    add_filters: bool = False,
) -> None:
    """Write multiple DataFrames to one consistently formatted Excel workbook.

    Each DataFrame gets its own worksheet with standard header styling
    (bold white on navy), thin borders, and best-fit column widths.
    Everything else is opt-in through the keyword arguments.

    Args:
        dfs (List[pd.DataFrame]): DataFrames to write; one per sheet.
        output_path (Union[str, Path]): Destination `.xlsx` file path
            (e.g. `"report.xlsx"` or a `Path`).
        sheet_names (Optional[List[str]], optional): Worksheet names. Missing entries become "Sheet1", "Sheet2"... Defaults to None.
        highlight_cols_list (Optional[List[Optional[List[str]]]], optional): Per sheet, the columns whose data cells get the
            light-blue highlight fill. Pass None or an empty list for sheets
            needing none. Defaults to None.
        freeze_header (bool, optional):  Keep the header row visible while scrolling. Defaults to False.
        wrap_threshold (int, optional): Column-width threshold (Excel character units)
            above which body cells receive `wrap_text=True`. Defaults to 40.
        as_table (bool, optional): Convert each data range into a native Excel table. Defaults to False.
        table_style (str, optional):  Excel Table style name, e.g. `"TableStyleMedium2"`.
            Only used when *as_table* is `True`. Defaults to `_DEFAULT_TABLE_STYLE`. Defaults to `_DEFAULT_TABLE_STYLE`.
        auto_fit_row_heights (bool, optional): Apply heuristic row-height adjustment
            for rows containing wrapped text. Defaults to False.
        show_grids (bool, optional): Show or hide the sheet grid-lines. Defaults to True.
        add_filters (bool, optional): Add AutoFilter dropdowns. Ignored when `as_table` is True,
            since tables include their own filters. Defaults to False.

    Raises:
        FileNotFoundError: If the destination folder does not exist.
        ValueError: If `highlight_cols_list` length does not match `dfs`.

    Example:
        >>> save_formatted_excel(
        ...     dfs=[df_due, df_overdue],
        ...     output_path="My_Report.xlsx",
        ...     sheet_names=["Due", "Overdue"],
        ...     highlight_cols_list=[["Protocol Name"], []],
        ...     freeze_header=True,
        ...     wrap_threshold=25,
        ...     as_table=True,
        ...     table_style="TableStyleMedium2",
        ...     auto_fit_row_heights=True,
        ...     show_grids=False,
        ... )
    """
    output_path = Path(output_path)

    # Fail early rather than after all the formatting work is done
    if not output_path.parent.exists():
        raise FileNotFoundError(f"The directory '{output_path.parent}' does not exist.")

    sheet_names = _ensure_sheet_names(dfs, sheet_names)
    highlight_cols_list = _validate_highlight_cols(dfs, highlight_cols_list)

    styles = _make_styles()

    # Filters overlay the header text, so widen the columns to compensate
    filter_padding = _FILTER_ICON_WIDTH if (as_table or add_filters) else 0.0

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for df, sheet_name, highlight_cols in zip(
                dfs, sheet_names, highlight_cols_list
            ):
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                ws.sheet_view.showGridLines = bool(show_grids)
                if freeze_header:
                    ws.freeze_panes = "A2"

                nrows, ncols = len(df), len(df.columns)

                # Nothing to format beyond the notice when there are no rows
                if nrows == 0:
                    _write_empty_sheet(ws, ncols, styles, wrap_threshold)
                    continue

                # 1. Headers, column widths, and text wrapping
                wrapped_cols = _format_headers_and_widths(
                    ws, df, styles, wrap_threshold, filter_padding
                )

                # 2. Borders and highlighted columns
                _apply_borders(ws, nrows, ncols, styles)
                _highlight_columns(ws, nrows, highlight_cols, styles)

                # 3. Row heights for the wrapped columns
                if auto_fit_row_heights and wrapped_cols:
                    _apply_wrap_row_heights(
                        ws, min_row=2, max_row=nrows + 1, wrap_cols=wrapped_cols
                    )

                # 4. Native Excel table, or plain AutoFilter
                _add_table_or_filter(
                    ws,
                    writer.book,
                    sheet_name,
                    nrows,
                    ncols,
                    as_table,
                    table_style,
                    add_filters,
                )

        print(f"\nFile saved and formatted: {output_path}")
        for name, df in zip(sheet_names, dfs):
            print(f"   └─ Sheet '{name}': {df.shape[0]} rows x {df.shape[1]} cols")

    except Exception as exc:
        print(f"Failed to save '{output_path}': {exc}")
        raise
